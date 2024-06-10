from typing import Callable, List
from Algorithms.Algorithm import Algorithm
from Task import Task
from Sim import Sim
import numpy
import os
import random
import openpyxl
from openpyxl import Workbook
from Algorithms.GreedyHeuristics import (
    MaxRuntimeFirst,
    MinRuntimeFirst,
    OutDegreesFirst,
    OutDegreesLast,
)

FILE_TO_INDEX = {}


class Threshold:
    def __init__(self, tasks: list["Task"]) -> None:
        self.tasks: list["Task"] = tasks

    def find_thresholds(
        self,
        recursion_depth: int,
    ):
        self._find_thresholds()

    def _find_thresholds(
        self,
        recursion_depth: int,
        task_list: List[Task],
        value_extractor: Callable[[Task], float],
    ) -> List[float]:
        n = len(task_list)
        if recursion_depth == 0 or n <= 1:
            return []
        # Use the value_extractor to get the threshold value
        thresh_value = value_extractor(task_list[n // 2])
        return (
            [thresh_value]
            + self.find_thresholds(
                recursion_depth - 1, task_list[: n // 2], value_extractor
            )
            + self.find_thresholds(
                recursion_depth - 1, task_list[n // 2 :], value_extractor
            )
        )

    def get_random_files(self, directory, num):
        all_files = os.listdir(directory)

        # filter out directories, only keep files
        files = [f for f in all_files if os.path.isfile(os.path.join(directory, f))]

        return random.sample(files, num)

    def test_thresholds(self, algorithm: Algorithm, thresholds, num_of_files=5):
        # rows = priority thresholds
        # columns = priority rates
        # cells = average time for a certain priority rate and threshold
        rates = [0.5, 0.6, 0.7, 0.8, 0.9, 1]
        runtimes = numpy.zeros((len(thresholds), len(rates)))  # cells
        rand_files = self.get_random_files("Parser/Data/parsed", num_of_files)

        # sum all runs in the cells and calculate average at the end
        print("finding thresholds...")
        for row, threshold in enumerate(thresholds):
            for column, rate in enumerate(rates):
                for file in rand_files:
                    sim = Sim()
                    sim.read_data(f"Parser/Data/parsed/{file}")
                    sim.start(
                        algorithm(
                            sim.tasks,
                            sim.processors,
                            sim.tasks,
                            False,
                            True,
                            threshold,
                            rate,
                        ),
                        illustration=False,
                    )
                    runtimes[row, column] += sim.final_end_time

        runtimes /= len(rand_files)
        return thresholds, rates, runtimes

    def check_find_threshold(self, attribute):
        min_val = getattr(self.tasks[0], attribute)
        max_val = getattr(self.tasks[-1], attribute)
        ranges = (
            [min_val]
            + sorted(self.find_thresholds(3, self.tasks, attribute))
            + [max_val]
        )

        res = {}
        for r in ranges:
            res[r] = 0

        for task in self.tasks:
            for i in range(1, len(ranges)):
                if ranges[i - 1] <= getattr(task, attribute) <= ranges[i]:
                    res[ranges[i - 1]] += 1
                    break
        return res
    
    def init_dictionary():
        files = os.listdir('Parser/Data/parsed')
        files_dict = {i+2: file for i, file in enumerate(files)}
        FILE_TO_INDEX = files_dict
        
    def init_sheet(self,
                   output_file: str,
                   algorithm: Algorithm,
                   thresholds: list['float']):
        if FILE_TO_INDEX == {}:
            self.init_dictionary()
        workbook: Workbook = openpyxl.load_workbook(output_file)
        sheet = None
        if algorithm.__qualname__ in workbook.sheetnames:
            sheet = workbook[algorithm.__qualname__]
            self.clear_sheet(sheet)
        else:
            workbook.create_sheet(algorithm.__qualname__)
            sheet = workbook[algorithm.__qualname__]
        for i, threshold in enumerate(thresholds):
            sheet.cell(1, i+2).value = threshold
        for key in FILE_TO_INDEX.keys:
            sheet.cell(key, 1).value = FILE_TO_INDEX[key]
        
        workbook.save(output_file)
            
    # writes the results to an excel spreadsheet
    def write_results(
        self,
        output_file: str,
        input_file: str,
        algorithm: Algorithm,
        threshold: float,
        runtime: float,
    ):
        workbook: Workbook = openpyxl.load_workbook(output_file)

        if algorithm.__qualname__ in workbook.sheetnames:
            sheet = workbook[algorithm.__qualname__]
        else:
            raise Exception("init the sheet before write result (use init_sheet())")

        #self.clear_sheet(sheet)
        def find_threshold_column(threshold):
            i = 1
            while i<11:
                if str(sheet.cell(0, i).value) == str(threshold):
                    return i 
            return -1

        sheet.cell(row=FILE_TO_INDEX[input_file], column=find_threshold_column(threshold)).value = runtime
        workbook.save(output_file)

    def clear_sheet(self, sheet):
        for row in sheet.iter_rows():
            for cell in row:
                cell.value = None
                cell.style = "Normal"


def main():
    sim = Sim()
    sim.read_data("Parser/Data/parsed/gsf.-00001.prof.json")
    print(f"len of tasks:  {len(sim.tasks)}")
    th = Threshold(sorted(sim.tasks, key=lambda x: x.duration))
    thresholds = th.find_thresholds(3, sim.tasks, "duration")
    algorithms = [MinRuntimeFirst, MaxRuntimeFirst, OutDegreesFirst, OutDegreesLast]
    greedy_runtime = 1 # TODO get the runtime of greedy algo
    for algorithm in algorithms:
        thresholds, rates, runtimes = th.test_thresholds(
            algorithm,
            thresholds,
            5,
        )
        # rates = [0.1,0.2,0.3,0.4]
        # thresholds=[234,5698,123]
        # runtimes = numpy.random.random((len(thresholds), len(rates)))
        th.init_sheet("Result.xlsx", algorithm=algorithm, thresholds=thresholds)
        for threshold in thresholds:
            for filename in os.listdir("Parser/Data/parsed"):
                runtime, _ = 1 #TODO get the runtime value 
                th.write_results("Results.xlsx",
                                 input_file=filename,
                                 algorithm=algorithm,
                                 threshold=threshold,
                                 runtime=runtime/greedy_runtime)


if __name__ == "__main__":
    main()
