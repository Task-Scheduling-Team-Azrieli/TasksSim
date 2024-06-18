from openpyxl import Workbook
import openpyxl
import random
from Task import Task
from Processor import Processor
from Algorithms.Algorithm import Algorithm
from Algorithms.Greedy import Greedy, MobileyeGreedy
from Algorithms.GreedyHeuristics import (
    OutDegreesFirst,
    MinRuntimeFirst,
    MaxRuntimeFirst,
    FromCriticalPath,
    OutDegreesLast,
)
from queue import PriorityQueue
from TimeLineIlustration import TimeLineIlustartion
from typing import List, Tuple, Dict
import json
import os

# for writing the results to excel
FILE_TO_INDEX = {}


class Sim:
    def __init__(self):
        self.tasks: List["Task"] = []
        self.processors: List["Processor"] = []

        self.algorithm = None

        self.total_time = 0
        self.final_end_time = 0

    def read_data(self, file_path: str):
        """reads data from a json file to objects (tasks, processors) in this class

        Args:
            file_path (str): full path to the file to read the data from
        """
        input_file = open(file_path)
        data = json.load(input_file)

        tasks_json = data["Tasks"]
        processors_json: List[str] = data["Processors"]

        # to connect a task name to its object, improves efficiency
        # when adding in and out degrees
        temp: Dict[str, Task] = {}

        # create all tasks
        for task_name in tasks_json:
            task_info = tasks_json[task_name]
            task = Task(
                task_name,
                task_info["duration"],
                task_info["processor_type"],
                task_info["priority"],
                [],
                [],
            )
            self.tasks.append(task)
            temp[task_name] = task

        # add in and out degrees to the tasks
        for task_name in tasks_json:
            # add out-degrees
            task_info = tasks_json[task_name]

            blocking = [
                temp[task_name_blocking] for task_name_blocking in task_info["blocking"]
            ]
            temp[task_name].blocking = blocking

            # add in-degrees
            for task in temp[task_name].blocking:
                task.blocked_by.append(temp[task_name])

        # read processors
        for processor_str in processors_json:
            processor_name = processor_str.split(":")[0]
            processor_type = int(processor_str.split(":")[1])
            processor = Processor(processor_name, processor_type)

            if processor not in self.processors:
                self.processors.append(processor)

        self.timeLineIlustartor = TimeLineIlustartion(self.processors)

    def start(
        self,
        algorithm: Algorithm,
        illustration=False,
    ):
        """starts the simulation, matches tasks for processors

        Args:
            algorithm (Algorithm): the algorithm to apply the order of tasks
            illustration (bool, optional): adds an illustration to the sim. Defaults to False.

        Returns:
            (float, int): total duration of all tasks, final end time of all tasks
        """
        self.algorithm = algorithm
        working_processors = []
        idle_processors: List["Processor"] = self.processors.copy()

        current_tasks: PriorityQueue[Tuple[float, Task]] = PriorityQueue()
        ready_tasks = [task for task in self.tasks if task.is_ready()]

        # for offline algorithms
        if algorithm.offline:
            order = algorithm.calculate()

        def match_ready_tasks(current_time):
            algorithm.update_lists(idle_processors, ready_tasks, self.tasks)
            if algorithm.offline:
                ready_tasks_order = algorithm.decide(order)
            else:
                ready_tasks_order = algorithm.decide()
            for task in ready_tasks_order:
                for processor in self.processors:
                    if processor.type == task.processor_type and processor.idle:
                        # work on the task
                        processor.work_on_task(task)
                        task.end_time = current_time + task.duration

                        # update processor lists
                        working_processors.append(processor)
                        idle_processors.remove(processor)

                        # update task lists
                        current_tasks.put((task.end_time, task))
                        ready_tasks.remove(task)

                        # a task can only run on one processor
                        break

        # init - assign all the tasks you can
        match_ready_tasks(0)

        # main loop
        while len(self.tasks) > 0:
            # pop the first task to finish
            current_time, done_task = current_tasks.get()
            self.total_time += done_task.duration
            self.final_end_time = done_task.end_time

            # free the processor and update in-degrees
            processor = done_task.processed_by
            processor.task_finished(ready_tasks)

            # add to time line
            if illustration:
                self.timeLineIlustartor.add_to_timeline(
                    processor, done_task, current_time - done_task.duration
                )

            # update working/idle processors
            working_processors.remove(processor)
            idle_processors.append(processor)

            self.tasks.remove(done_task)

            match_ready_tasks(current_time)

        return self.total_time, self.final_end_time

    def show_illustration(self):
        self.timeLineIlustartor.show()

    def set_critical_path(self, critical_path):
        self.timeLineIlustartor.set_critical_path(critical_path)

    def __str__(self):
        return (
            f"{self.algorithm.__class__.__qualname__} End Time: {self.final_end_time}\n"
        )


def run_sim_once(
    algorithm: Algorithm,
    file_path: str,
    illustration=False,
    offline=False,
    is_mobileye: bool = False,
    is_critical: bool = False,
    thresholds: list["str | float"] = [],
    output_file: str = "Results.xlsx",
):
    sim = Sim()
    sim.read_data(file_path)

    if algorithm.__qualname__ == "FromCriticalPath" and len(thresholds) > 1:
        run_sim_once(
            algorithm,
            file_path,
            illustration,
            offline,
            is_mobileye,
            is_critical,
            ["Offline" if offline else "Online"],
            output_file,
        )
        return sim

    for threshold in thresholds:
        # run sim and get total time
        algorithm_instance = algorithm(
            sim.tasks,
            sim.processors,
            sim.tasks,
            offline,
            is_mobileye,
            is_critical,
            threshold,
        )
        _, total_time = sim.start(
            algorithm_instance,
            illustration=False,
        )
        # write to excel
        write_results(
            output_file,
            input_file=file_path,
            algorithm=algorithm_instance,
            threshold=threshold,
            runtime=total_time,
        )

        # init sim for the new threshold
        sim = Sim()
        sim.read_data(file_path)

    if illustration:
        critical_path_instance = FromCriticalPath(
            sim.tasks, sim.processors, sim.tasks, offline, is_mobileye
        )

        critical_path = critical_path_instance.calculate()
        critical_path = [task.name for task in critical_path]

        sim.set_critical_path(critical_path)
        sim.show_illustration()

    return sim

    # writes the results to an excel spreadsheet


def write_results(
    output_file: str,
    input_file: str,
    algorithm: Algorithm,
    threshold: float,
    runtime: float,
):
    input_file = input_file.split("/")[-1]
    workbook: Workbook = openpyxl.load_workbook(output_file)

    if algorithm.__class__.__name__ in workbook.sheetnames:
        sheet = workbook[algorithm.__class__.__name__]
    else:
        raise Exception("init the sheet before write result (use init_sheet())")

    def same_value(value1, value2):
        if type(value1) == str and type(value2) == str:
            return value1 == value2

        if type(value1) != type(value2):
            if type(value1) == str or type(value2) == str:
                return False

        return abs(value1 - value2) < 0.1 * value1

    def find_threshold_column(threshold):
        if not algorithm.is_mobileye:
            return 2
        i = 2
        while i < 13:
            if same_value(sheet.cell(1, i).value, threshold):
                return i
            i += 1
        return -1

    sheet.cell(
        row=FILE_TO_INDEX[input_file], column=find_threshold_column(threshold)
    ).value = runtime
    workbook.save(output_file)
    workbook.close()


def init_dictionary():
    global FILE_TO_INDEX
    files = os.listdir("Parser/Data/parsed")
    files_dict = {file.split("/")[-1]: i + 2 for i, file in enumerate(files)}
    FILE_TO_INDEX = files_dict


def init_sheet(output_file: str, algorithm_name: str, thresholds: list["float"]):
    workbook: Workbook = openpyxl.load_workbook(output_file)
    if "Sheet1" in workbook.sheetnames:
        del workbook["Sheet1"]
    sheet = None
    if algorithm_name in workbook.sheetnames:
        sheet = workbook[algorithm_name]
        clear_sheet(sheet)
    else:
        workbook.create_sheet(algorithm_name)
        sheet = workbook[algorithm_name]

    sheet.cell(1, 2).value = "Regular"
    for i, threshold in enumerate(thresholds):
        sheet.cell(1, i + 3).value = threshold

    for key in FILE_TO_INDEX.keys():
        sheet.cell(FILE_TO_INDEX[key], 1).value = key

    workbook.save(output_file)


def clear_sheet(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            cell.value = None
            cell.style = "Normal"


def run_sim_all(
    algorithm: Algorithm,
    folder_path: str,
    output_file: str,
    illustration=False,
    offline=False,
    is_mobileye: bool = False,
    is_critical: bool = False,
    thresholds: list["float"] = [],
):
    count = 0
    for filename in os.listdir(folder_path):
        run_sim_once(
            algorithm,
            f"{folder_path}/{filename}",
            False,
            offline,
            is_mobileye,
            is_critical,
            thresholds[algorithm.__qualname__],
        )
        print(f"done with {filename}")
        count += 1


def init_sheets_and_thresholds(output_file, num_rand_files=5):
    sim = Sim()
    random_files = random.sample(
        [
            "Parser/Data/parsed/" + file_name
            for file_name in os.listdir("Parser/Data/parsed")
        ],
        num_rand_files,
    )
    for random_file in random_files:
        sim.read_data(random_file)
    init_dictionary()
    algorithms: list[Algorithm] = [
        MinRuntimeFirst(
            sim.tasks, sim.processors, sim.tasks, offline=False, is_mobileye=True
        ),
        MaxRuntimeFirst(
            sim.tasks, sim.processors, sim.tasks, offline=False, is_mobileye=True
        ),
        OutDegreesFirst(
            sim.tasks, sim.processors, sim.tasks, offline=False, is_mobileye=True
        ),
        OutDegreesLast(
            sim.tasks, sim.processors, sim.tasks, offline=False, is_mobileye=True
        ),
    ]

    thresholds = {}
    # not mobileye
    thresholds["FromCriticalPath"] = ["Offline", "Online"]
    thresholds["Greedy"] = ["Regular"]
    init_sheet(output_file, "FromCriticalPath", thresholds["FromCriticalPath"])
    init_sheet(output_file, "Greedy", thresholds["Greedy"])

    # yes mobileye
    for algo in algorithms:
        thresholds[algo.__class__.__name__] = algo.find_thresholds(3)
        init_sheet(
            output_file, algo.__class__.__name__, thresholds[algo.__class__.__name__]
        )

    return thresholds


def main():
    output_file = "Results.xlsx"
    thresholds = init_sheets_and_thresholds(output_file)
"""
    # FromCriticalPath
    run_sim_all(
        FromCriticalPath,
        "Parser/Data/parsed",
        output_file,
        offline=False,
        is_mobileye=True,
        is_critical=False,
        thresholds=thresholds,
    )
    run_sim_all(
        FromCriticalPath,
        "Parser/Data/parsed",
        output_file,
        offline=True,
        is_mobileye=False,
        is_critical=False,
        thresholds=thresholds,
    )

    # Greedy
    run_sim_all(
        Greedy,
        "Parser/Data/parsed",
        output_file,
        is_mobileye=False,
        is_critical=False,
        thresholds=thresholds,
    )

    # MinRuntimeFirst
    run_sim_all(
        MinRuntimeFirst,
        "Parser/Data/parsed",
        output_file,
        is_mobileye=False,
        is_critical=False,
        thresholds=thresholds,
    )
    run_sim_all(
        MinRuntimeFirst,
        "Parser/Data/parsed",
        output_file,
        is_mobileye=True,
        is_critical=False,
        thresholds=thresholds,
    )
"""
    # MaxRuntimeFirst
    run_sim_all(
        MaxRuntimeFirst,
        "Parser/Data/parsed",
        output_file,
        is_mobileye=False,
        is_critical=False,
        thresholds=thresholds,
    )
    run_sim_all(
        MaxRuntimeFirst,
        "Parser/Data/parsed",
        output_file,
        is_mobileye=True,
        is_critical=False,
        thresholds=thresholds,
    )

    # OutDegreesLast
    run_sim_all(
        OutDegreesLast,
        "Parser/Data/parsed",
        output_file,
        is_mobileye=False,
        is_critical=False,
        thresholds=thresholds,
    )
    run_sim_all(
        OutDegreesLast,
        "Parser/Data/parsed",
        output_file,
        is_mobileye=True,
        is_critical=False,
        thresholds=thresholds,
    )

    # OutDegreesFirst
    run_sim_all(
        OutDegreesFirst,
        "Parser/Data/parsed",
        output_file,
        is_mobileye=False,
        is_critical=False,
        thresholds=thresholds,
    )
    run_sim_all(
        OutDegreesFirst,
        "Parser/Data/parsed",
        output_file,
        is_mobileye=True,
        is_critical=False,
        thresholds=thresholds,
    )


if __name__ == "__main__":
    main()
